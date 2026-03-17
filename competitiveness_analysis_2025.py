"""
홉아이엔씨 2025년도 회사경쟁력 분석 보고서
데이터 소스:
  - 2025 복식장부.xlsm  (손익계산서 / 재무상태표 / 2025 장부)
  - 품번별 매출이익_2602.xlsx
"""

import openpyxl
from collections import defaultdict

# ─────────────────────────────────────────────
# 데이터 로드
# ─────────────────────────────────────────────
def load_pl():
    wb = openpyxl.load_workbook('2025 복식장부.xlsm', read_only=True, keep_vba=True, data_only=True)
    ws = wb['(법인)손익계산서']
    data = {}
    for row in ws.iter_rows(values_only=True):
        if row[1] is not None and isinstance(row[1], int) and row[3] is not None:
            data[row[1]] = row[3]
        if row[6] is not None and isinstance(row[6], int) and row[7] is not None:
            data[row[6]] = row[7]
        if row[10] is not None and isinstance(row[10], int) and len(row) > 12 and row[12] is not None:
            data[row[10]] = row[12]
        if len(row) > 16 and row[15] is not None and isinstance(row[15], int) and row[16] is not None:
            data[row[15]] = row[16]
    return {k: round(v) for k, v in data.items() if isinstance(v, (int, float))}


def load_monthly():
    wb = openpyxl.load_workbook('2025 복식장부.xlsm', read_only=True, keep_vba=True, data_only=True)
    ws = wb['2025 장부']
    monthly = defaultdict(lambda: {'매출': 0, '외주': 0, '급여': 0, '이자': 0, '기타': 0})
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5:
            continue
        date = row[0]
        if not hasattr(date, 'month') or date.year != 2025:
            continue
        m = date.month
        for ca, va in [(20, 22), (24, 26)]:
            acc = str(row[ca]) if row[ca] else ''
            amt = float(row[va]) if isinstance(row[va], (int, float)) else 0
            if '매출' in acc:
                monthly[m]['매출'] += amt
        for cd, vd in [(12, 14), (16, 18)]:
            acc = str(row[cd]) if row[cd] else ''
            amt = float(row[vd]) if isinstance(row[vd], (int, float)) else 0
            if '외주' in acc or '용역' in acc:
                monthly[m]['외주'] += amt
            elif '급여' in acc or '일용' in acc:
                monthly[m]['급여'] += amt
            elif '이자' in acc:
                monthly[m]['이자'] += amt
    return monthly


def load_clients():
    wb = openpyxl.load_workbook('2025 복식장부.xlsm', read_only=True, keep_vba=True, data_only=True)
    ws = wb['2025 장부']
    client = defaultdict(float)
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5:
            continue
        date = row[0]
        if not hasattr(date, 'month') or date.year != 2025:
            continue
        c = str(row[11]).strip() if row[11] else '(불명)'
        for ca, va in [(20, 22), (24, 26)]:
            acc = str(row[ca]) if row[ca] else ''
            amt = float(row[va]) if isinstance(row[va], (int, float)) else 0
            if '매출' in acc and amt > 0:
                client[c] += amt
    return client


def load_product():
    wb = openpyxl.load_workbook('품번별 매출이익_2602.xlsx', read_only=True, data_only=True)
    ws = wb.active
    group  = defaultdict(lambda: {'매출': 0, '원료비': 0, '가공비': 0, '이익': 0, '건수': 0})
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 4:
            continue
        g      = row[5]
        rev    = float(row[22]) if isinstance(row[22], (int, float)) else 0
        cogs   = float(row[16]) if isinstance(row[16], (int, float)) else 0
        proc   = float(row[18]) if isinstance(row[18], (int, float)) else 0
        profit = float(row[25]) if isinstance(row[25], (int, float)) else 0
        if rev == 0 and profit == 0:
            continue
        if g:
            group[g]['매출']   += rev
            group[g]['원료비'] += cogs
            group[g]['가공비'] += proc
            group[g]['이익']   += profit
            group[g]['건수']   += 1
    return group


# ─────────────────────────────────────────────
# 출력 유틸
# ─────────────────────────────────────────────
SEP = "=" * 70
sep = "-" * 70

def w(n):
    return f"{int(n):,}"

def m(n):
    """백만원 단위"""
    return f"{n/1e6:,.1f}백만"

def pct(a, b):
    return f"{a/b*100:.1f}%" if b else "-"

def bar(ratio, width=20):
    filled = int(ratio * width)
    return "█" * filled + "░" * (width - filled)


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    pl      = load_pl()
    monthly = load_monthly()
    clients = load_clients()
    product = load_product()

    rev        = pl.get(1,  0)   # 매출액
    cogs       = pl.get(35, 0)   # 매출원가
    gross      = pl.get(66, 0)   # 매출총이익
    sga        = pl.get(67, 0)   # 판관비
    op_inc     = pl.get(129, 0)  # 영업이익
    non_rev    = pl.get(130, 0)  # 영업외수익
    non_exp    = pl.get(179, 0)  # 영업외비용
    int_exp    = pl.get(180, 0)  # 이자비용
    net        = pl.get(219, 0)  # 당기순이익
    outsrc     = pl.get(121, 0)  # 외주용역비
    salary     = pl.get(68, 0)   # 급여

    # 티에이케이 계열 통합
    tak_rev = sum(v for k, v in clients.items() if '티에이케이' in k)
    silong  = sum(v for k, v in clients.items() if '실론' in k)
    himel   = sum(v for k, v in clients.items() if '하이멜' in k or '문정환' in k)
    total_client_rev = sum(clients.values())
    top3_rev = tak_rev + silong + himel

    # 월별
    months_rev = [monthly[m]['매출'] for m in range(1, 13)]
    max_m = max(range(1, 13), key=lambda x: monthly[x]['매출'])
    min_m = min(range(1, 13), key=lambda x: monthly[x]['매출'])
    avg_m = sum(months_rev) / 12

    print()
    print(SEP)
    print("      홉아이엔씨 (Hope I&C)  2025년도 회사경쟁력 분석 보고서")
    print(SEP)

    # ── 1. 수익성 ──────────────────────────────────────────
    print()
    print("【 1. 수익성 분석 】")
    print(sep)
    metrics = [
        ("매출총이익률",  gross / rev,    "  (원가 효율성 ↑ 수록 좋음)"),
        ("영업이익률",    op_inc / rev,   "  (핵심 영업 효율)"),
        ("당기순이익률",  net / rev,      "  (최종 수익성)"),
        ("이자비용/영업이익", int_exp / op_inc, "  (금융비용 부담)"),
        ("외주비/매출",   outsrc / rev,   "  (원가 구조의 핵심)"),
    ]
    for name, ratio, note in metrics:
        b = bar(min(ratio, 1.0))
        sign = "▲ 주의" if name == "이자비용/영업이익" and ratio > 0.4 else ""
        sign = "▲ 주의" if name == "외주비/매출" and ratio > 0.5 else sign
        print(f"  {name:<20}  {ratio*100:6.1f}%  [{b}] {note}  {sign}")

    print()
    print(f"  ┌─ 수익성 요약 ───────────────────────────────────────────┐")
    print(f"  │  매출액       {w(rev):>20}원                          │")
    print(f"  │  영업이익     {w(op_inc):>20}원  (이익률 {pct(op_inc,rev)})          │")
    print(f"  │  당기순이익   {w(net):>20}원  (이익률 {pct(net,rev)})          │")
    print(f"  │  이자비용이 영업이익의 {pct(int_exp, op_inc)}를 소진 → 금융 부담 큼        │")
    print(f"  └─────────────────────────────────────────────────────────┘")

    # ── 2. 매출 구조 ────────────────────────────────────────
    print()
    print("【 2. 매출 구조 분석 】")
    print(sep)
    dom = pl.get(6, 0)
    exp = pl.get(7, 0)
    print(f"  국내 매출   {w(dom):>18}원  ({pct(dom, rev)})  [{bar(dom/rev)}]")
    print(f"  수출 매출   {w(exp):>18}원  ({pct(exp, rev)})  [{bar(exp/rev)}]")
    print()
    print(f"  → 국내/수출 비중 약 6:4  ·  수출 비중 확대가 성장 동력")

    # ── 3. 거래처 집중도 ─────────────────────────────────────
    print()
    print("【 3. 거래처 집중도 분석 (매출 편중 리스크) 】")
    print(sep)
    client_sorted = sorted(
        [(k, v) for k, v in clients.items() if v > 5_000_000],
        key=lambda x: -x[1]
    )
    # 그룹화
    groups = [
        ("티에이케이 계열", tak_rev),
        ("실론 계열",       silong),
        ("하이멜 계열",     himel),
    ]
    for name, gr in groups:
        b = bar(gr / total_client_rev)
        print(f"  {name:<20} {w(gr):>16}원  ({pct(gr, total_client_rev)})  [{b}]")
    others = total_client_rev - tak_rev - silong - himel
    print(f"  {'기타 거래처':<20} {w(others):>16}원  ({pct(others, total_client_rev)})  [{bar(others/total_client_rev)}]")
    print()
    conc = pct(top3_rev, total_client_rev)
    print(f"  ★ 상위 3개 그룹 매출 집중도: {conc}  {'▲ 집중도 매우 높음 – 거래처 다변화 필요' if top3_rev/total_client_rev > 0.6 else '적정 수준'}")

    # ── 4. 제품군별 수익성 ──────────────────────────────────
    print()
    print("【 4. 제품군별 수익성 (2026년 초 기준) 】")
    print(sep)
    print(f"  {'제품군':<12}  {'매출':>12}  {'이익':>10}  {'이익률':>7}  {'이익률 BAR'}")
    print(f"  {'-'*12}  {'-'*12}  {'-'*10}  {'-'*7}  {'-'*22}")
    prod_sorted = sorted(product.items(), key=lambda x: -x[1]['이익'])
    for g, v in prod_sorted:
        pm = v['이익'] / v['매출'] if v['매출'] else 0
        b  = bar(max(pm, 0))
        flag = "  ★ 고수익" if pm > 0.35 else ("  ▲ 저수익" if pm < 0.10 else "")
        print(f"  {str(g):<12}  {w(v['매출']):>12}  {w(v['이익']):>10}  {pct(v['이익'],v['매출']):>7}  [{b}]{flag}")
    print()
    cleaner_margin = product.get('cleaner', {}).get('이익', 0)
    cleaner_rev    = product.get('cleaner', {}).get('매출', 1)
    print(f"  → cleaner 제품군 이익률 {pct(cleaner_margin, cleaner_rev)} 저조 – 원가구조 개선 필요")
    print(f"  → tape 제품군 이익률 39.8%로 최우선 확대 전략 권고")

    # ── 5. 비용 구조 ────────────────────────────────────────
    print()
    print("【 5. 비용 구조 분석 】")
    print(sep)
    cost_items = [
        ("외주용역비",      outsrc,          "  ← 비용의 核心 항목"),
        ("급여",            salary,          ""),
        ("기타판관비",      pl.get(124, 0),  ""),
        ("지급수수료",      pl.get(104, 0),  ""),
        ("차량유지비",      pl.get(93, 0),   ""),
        ("여비교통비",      pl.get(80, 0),   ""),
        ("연구비",          pl.get(94, 0),   ""),
        ("수출입제비용",    pl.get(119, 0),  ""),
        ("임차료",          pl.get(81, 0),   ""),
        ("접대비",          pl.get(85, 0),   ""),
    ]
    total_cost = sga
    for name, amt, note in cost_items:
        if amt == 0:
            continue
        b = bar(amt / total_cost)
        print(f"  {name:<14}  {w(amt):>14}원  ({pct(amt, total_cost)})  [{b}]{note}")

    # ── 6. 월별 매출 추이 ──────────────────────────────────
    print()
    print("【 6. 월별 매출 추이 (계절성·변동성) 】")
    print(sep)
    for month in range(1, 13):
        rev_m = monthly[month]['매출']
        b = bar(rev_m / max(months_rev))
        flag = "◀ 최고" if month == max_m else ("◀ 최저" if month == min_m else "")
        print(f"  {month:2d}월  {w(rev_m):>14}원  [{b}] {flag}")
    print()
    volatility = (max(months_rev) - min(months_rev)) / avg_m * 100
    print(f"  평균 월매출: {m(avg_m)}  |  변동성(최대-최소/평균): {volatility:.0f}%")
    print(f"  → {max_m}월 최고·{min_m}월 최저  |  계절 편차 큼 → 안정적 수주 파이프라인 필요")

    # ── 7. SWOT 종합 ────────────────────────────────────────
    print()
    print("【 7. SWOT 종합 분석 】")
    print(SEP)
    swot = {
        "강점 (Strengths)": [
            f"높은 매출총이익률 {pct(gross,rev)} – 부가가치 높은 사업 모델",
            f"수출 비중 40.5% – 글로벌 시장 진출 기반 확보",
            f"tape 제품군 이익률 39.8% – 핵심 고수익 라인 보유",
            f"다양한 섬유 제품군 (tape/cleaner/blind/back) 포트폴리오",
        ],
        "약점 (Weaknesses)": [
            f"영업이익률 {pct(op_inc,rev)} / 당기순이익률 {pct(net,rev)} – 최종 수익성 낮음",
            f"외주용역비가 판관비의 {pct(outsrc,sga)} 차지 – 외주 의존도 과다",
            f"이자비용이 영업이익의 {pct(int_exp,op_inc)} 소진 – 금융비용 부담",
            f"cleaner 제품군 이익률 6.7% – 저마진 구조",
        ],
        "기회 (Opportunities)": [
            "수출 매출 40.5% → 글로벌 섬유 시장 확대 여지",
            "tape/PTFE 등 기능성 소재 수요 성장 트렌드",
            "원사 직판매(이익률 78.6%) 등 고마진 채널 확대 가능",
            "연구개발비 투자(6.6백만) → 신제품 개발 기반 있음",
        ],
        "위협 (Threats)": [
            f"거래처 집중도 상위3 {pct(top3_rev,total_client_rev)} – 특정 고객 이탈 시 큰 타격",
            "월매출 변동성 최대-최소 2.3배 – 수주 불안정",
            "원재료·가공비 상승 시 cleaner 라인 적자 전환 위험",
            "이자비용 부담 지속 시 당기순이익 적자 전환 가능",
        ],
    }
    for section, points in swot.items():
        print(f"\n  ▶ {section}")
        for p in points:
            print(f"     • {p}")

    # ── 8. 전략 제언 ────────────────────────────────────────
    print()
    print("【 8. 전략 제언 】")
    print(SEP)
    recs = [
        ("수익성 개선",   [
            "외주비 내재화 또는 협력사 단가 재협상 → 영업이익률 5%↑ 목표",
            "cleaner 라인 원가분석 후 단가 인상 또는 비중 축소 검토",
            "tape·PTFE 고수익 제품 비중 확대 집중",
        ]),
        ("거래처 다변화", [
            "티에이케이·실론·하이멜 3사 집중도 67% → 신규 바이어 발굴",
            "수출 국가 다변화로 단일 지역 의존 리스크 분산",
        ]),
        ("재무 안정성",   [
            f"이자비용 {m(int_exp)} – 단기 차입금 상환 우선순위 설정",
            "매출채권 회수 가속화로 운전자본 개선",
        ]),
        ("성장 전략",     [
            "원사 직판매(이익률 78.6%) 채널 적극 확대",
            "PTFE/기능성 소재 라인 – 글로벌 스포츠·아웃도어 시장 공략",
            "수출 비중 40.5% → 2026년 50% 이상 목표 설정",
        ]),
    ]
    for title, items in recs:
        print(f"\n  [{title}]")
        for it in items:
            print(f"    ✓ {it}")

    print()
    print(SEP)
    print("  ※ 본 분석은 2025 복식장부.xlsm 및 품번별 매출이익_2602.xlsx 기준입니다.")
    print(SEP)
    print()


if __name__ == '__main__':
    main()
